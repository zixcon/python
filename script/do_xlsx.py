import openpyxl
from collections import Counter

wb = openpyxl.load_workbook(filename=r'./数据(2017.12.29)(1).xlsx')

ws1 = wb.get_sheet_by_name("Squirrel SQL Export")

print(ws1.cell(row=1, column=1).value)
print(ws1.cell(row=1, column=3).value)
print(ws1.cell(row=1, column=4).value)
print(ws1.cell(row=1, column=5).value)

case_id_str = ''
case_id_arr = []
total = 0

f = open('sql.txt', 'a')

# sql
sql = "update table_name_1 set person_id = (select id from table_name_2 where real_name= '%s' and team_name= '%s' ) ,person_name= '%s', team_name= '%s', status=1 where id = %d and type=1  and amount=0 and status !=-1;"
for rows in range(2, 5911):
    col1 = ws1.cell(row=rows, column=1).value
    col3 = ws1.cell(row=rows, column=3).value
    col4 = ws1.cell(row=rows, column=4).value
    col5 = ws1.cell(row=rows, column=5).value
    sql_str = sql % (col4, col5, col4, col5, col1)
    f.writelines(sql_str)
    # print(sql_str)
    case_id_str = case_id_str + str(col1) + ','
    case_id_arr.append(col1)
    total += 1
f.close()
# print(case_id_str)
print(case_id_arr)
print(total)
print(Counter(case_id_arr))


# for sheet_name in wb.get_named_ranges():
#     sheet=wb.get_sheet_by_name(sheet_name)
#     for i in range(1,sheet.max_row+1):
#         print(sheet.cell(row=i,column=1).value)

